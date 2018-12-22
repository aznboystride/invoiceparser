import pytesseract
import openpyxl
from PIL import Image
from datetime import datetime
from copy import copy

class ImageReader(object):

        def __init__(self, path, psm=None):
                self.image = Image.open(path)
                if psm == None:
                        self.string = pytesseract.image_to_string(self.image) # ,config='--psm 6')
                else:
                        self.string = pytesseract.image_to_string(self.image, config='--psm {}'.format(psm))


        def getStringFromImage(self):
                return self.string

        def getSanitizeStringFromImage(self):
                dirtyString = self.string
                cleanString = ''
                for line in dirtyString.split('\n'):
                        if '/' in line.lower() or 'job' in line.lower():
                                cleanString += (line + '\n')

                badChars = list()
                for c in cleanString:
                        if ord(c) not in range(128):
                                badChars.append(c)
                for b in badChars:
                        cleanString = cleanString.replace(b, '')
                lines = cleanString.split('\n')
                while '' in lines:
                        lines.remove('')
                return lines


class InvoiceWriter(object):

        dateColumn = 1
        jobColumn = 2
        trackIDColumn = 1
        fromColumn = 3
        toColumn = 4
        amtColumn = 7

        def __init__(self, path):
                self.path = path
                self.wb = openpyxl.load_workbook(path)
                self.sheet = self.wb['Sales Invoice']

                self.dateFormat = self.sheet['a17'].number_format
                self.dateFont = copy(self.sheet['a17'].font)
                self.dateAlignment = copy(self.sheet['a17'].alignment)

                self.trackIDFormat = copy(self.sheet['a18'].number_format)
                self.trackIDFont = copy(self.sheet['a18'].font)
                self.trackIDAlignment = copy(self.sheet['a18'].alignment)

                self.fromFormat = copy(self.sheet['c18'].number_format)
                self.fromFont = copy(self.sheet['c18'].font)
                self.fromAlignment = copy(self.sheet['c18'].alignment)

                self.toFormat = copy(self.sheet['c18'].number_format)
                self.toFont = copy(self.sheet['c18'].font)
                self.toAlignment = copy(self.sheet['c18'].alignment)

                self.amtFormat = copy(self.sheet['g18'].number_format)
                self.amtFont = copy(self.sheet['g18'].font)
                self.amtAlignment = copy(self.sheet['g18'].alignment)

                self.jobFormat = copy(self.sheet['b18'].number_format)
                self.jobFont = copy(self.sheet['b18'].font)
                self.jobAlignment = copy(self.sheet['b18'].alignment)

        def deleteBlankRows(self, totalrow):
                for row in range(16, self.sheet.max_row + 1):
                        if self.sheet.cell(row=row, column=1).value is None:
                                self.sheet.delete_rows(row, totalrow-row)
                                self.sheet.merge_cells(start_row=row+5, start_column=1, end_row=row+6, end_column=7)
                                break

        def writeJob(self, row, date, trackID, job, fr, to, price):

               # day = int(date[date.find('/') + 1 : date.find('/', date.find('/') + 1)])
               # month = int(date[:date.find('/')])
               # year = int(date[date.find('/', date.find('/') + 1) + 1:])

                self.sheet.cell(row=row+1, column=self.jobColumn).value = job
                self.sheet.cell(row=row+1, column=self.trackIDColumn).value = trackID
                self.sheet.cell(row=row, column=self.dateColumn).value = date
                self.sheet.cell(row=row+1, column=self.fromColumn).value = fr
                self.sheet.cell(row=row+1, column=self.toColumn).value = to
                self.sheet.cell(row=row+1, column=self.amtColumn).value = price

                self.setDateFormats(row)
                self.setTrackIDFormats(row)
                self.setFromFormats(row)
                self.setToFormats(row)
                self.setAmtFormats(row)
                self.setJobFormats(row)

        def writeTotal(self, amt, row):
            self.sheet['g{}'.format(row)].value = amt
            self.sheet['g{}'.format(row+1)].value = '{:.2f}'.format(-1 * float(amt) / 10)
            self.sheet['g{}'.format(row+2)].value = '{:.2f}'.format(float(amt) * .9)

        def writeInvoiceNumber(self, num):
                self.sheet['e5'].value = num

        def writeInvoiceDateCreation(self, date):
                self.sheet['e6'].value = date

        def setDateFormats(self, row):
                self.sheet.cell(row=row, column=self.dateColumn).number_format = 'd-mmm-yy'
                self.sheet.cell(row=row, column=self.dateColumn).font = self.dateFont
                self.sheet.cell(row=row, column=self.dateColumn).alignment = self.dateAlignment

        def setTrackIDFormats(self, row):
                self.sheet.cell(row=row+1, column=self.trackIDColumn).number_format = self.trackIDFormat
                self.sheet.cell(row=row+1, column=self.trackIDColumn).font = self.trackIDFont
                self.sheet.cell(row=row+1, column=self.trackIDColumn).alignment = self.trackIDAlignment

        def setFromFormats(self, row):
                self.sheet.cell(row=row+1, column=self.fromColumn).number_format = self.fromFormat
                self.sheet.cell(row=row+1, column=self.fromColumn).font = self.fromFont
                self.sheet.cell(row=row+1, column=self.fromColumn).alignment = self.fromAlignment

        def setToFormats(self, row):
                self.sheet.cell(row=row+1, column=self.toColumn).number_format = self.toFormat
                self.sheet.cell(row=row+1, column=self.toColumn).font = self.toFont
                self.sheet.cell(row=row+1, column=self.toColumn).alignment = self.toAlignment

        def setAmtFormats(self, row):
                self.sheet.cell(row=row+1, column=self.amtColumn).number_format = self.amtFormat
                self.sheet.cell(row=row+1, column=self.amtColumn).font = self.amtFont
                self.sheet.cell(row=row+1, column=self.amtColumn).alignment = self.amtAlignment

        def setJobFormats(self, row):
                self.sheet.cell(row=row+1, column=self.jobColumn).number_format = self.jobFormat
                self.sheet.cell(row=row+1, column=self.jobColumn).font = self.jobFont
                self.sheet.cell(row=row+1, column=self.jobColumn).alignment = self.jobAlignment


        def finalize(self, path):
                self.wb.save(path)


class InvoiceReader(object):

        dateColumn = 1
        jobColumn = 4
        trackIDColumn = 3
        fromColumn = 7
        toColumn = 8
        amtColumn = 9

        def __init__(self, path):
                self.path = path
                self.wb = openpyxl.load_workbook(path)
                self.sheet = self.wb['Sheet 1']

        def getInfoDictionary(self, job):
                info = dict()
                info['job'] = self.getJobGivenJobNum(job)
                info['trackID'] = self.getTrackIDGivenJobNum(job)
                info['date'] = self.getDateGivenJobNum(job)
                info['from'] = self.getFromGivenJobNum(job)
                info['to'] = self.getToGivenJobNum(job)
                info['amt'] = self.getAmtGivenJobNum(job)
                return info


        def getDateGivenJobNum(self, job):
                for row in range(3, self.sheet.max_row + 1):
                        try:
                                if str(job) in self.sheet.cell(row=row, column=self.jobColumn).value:
                                        return self.sheet.cell(row=row, column=self.dateColumn).value
                        except:
                                break
        def getTrackIDGivenJobNum(self, job):
                for row in range(3, self.sheet.max_row + 1):
                        try:
                                if str(job) in self.sheet.cell(row=row, column=self.jobColumn).value:
                                        return self.sheet.cell(row=row, column=self.trackIDColumn).value
                        except:
                                break
        def getJobGivenJobNum(self, job):
                for row in range(3, self.sheet.max_row + 1):
                        try:
                                if str(job) in self.sheet.cell(row=row, column=self.jobColumn).value:
                                        return self.sheet.cell(row=row, column=self.jobColumn).value
                        except:
                                break

        def getFromGivenJobNum(self, job):
                for row in range(3, self.sheet.max_row + 1):
                        try:
                                if str(job) in self.sheet.cell(row=row, column=self.jobColumn).value:
                                        return self.sheet.cell(row=row, column=self.fromColumn).value
                        except:
                                break
        def getToGivenJobNum(self, job):
                for row in range(3, self.sheet.max_row + 1):
                        try:
                                if str(job) in self.sheet.cell(row=row, column=self.jobColumn).value:
                                        return self.sheet.cell(row=row, column=self.toColumn).value
                        except:
                                break
        def getAmtGivenJobNum(self, job):
                for row in range(3, self.sheet.max_row + 1):
                        try:
                                if str(job) in self.sheet.cell(row=row, column=self.jobColumn).value:
                                        return self.sheet.cell(row=row, column=self.amtColumn).value
                        except:
                                break
